from tkinter import *
import os
import openpyxl
wb= openpyxl.load_workbook('data1/student_marks.xlsx')

sheetname = wb.get_sheet_names()

sheet=wb.get_sheet_by_name(sheetname[0])


opt = [
    "java", "python", "c#", "c", "c++", "networking", "operating system", "web technology", "software engineering",
    "internet of things", "data mining", "software project management", "data science", "english", "data structure",
    "artificial intelligence", "computer grapfics", "mathematics",
    "chemistry", "mobile computing", "natural language processing", "android development", "aws", "cyber security",
    "machine learning", "embeded system", "physics",
    "digital electronics", "image processing", "cloud computing", "linux", "oracle", "rdbms", "dbms", "statistics"
]

window = Tk()
window.title("COURSE RECOMMENDATION SYSTEM")
window.geometry('870x600')
lbA = Label(window, text="WELCOME  TO  COURSE  RECOMMENDATION  SYSTEM", fg='blue', font=('Arial', 20)).place(x=60,y=40)

window['bg'] = '#5d8a82'

data = StringVar()
data.set("CHOOSE")
drop = OptionMenu(window, data, *opt).place(x=470,y=240,width=230)



def IsScore(sheet,r,c):
    if sheet.cell(row = r + 1, column=c).value != None:
        return True
    else:
        return False

def scr_sub(sheet,u,i):
    return sheet.cell(row = u+1,column = i).value


def getAvgGrades(sheet,student):    # calculate avg grades for student
    sum = 0
    cnt = 0
    for i in range (2,sheet.max_column-1):
        if IsScore(sheet,student,i):
            sum = sum + int(scr_sub(sheet,student,i))
            cnt = cnt + 1
    if cnt != 0:
        return round(float(sum/cnt),3)
    else:
        return 0


def similarity(c1, c2, sheet):     #similarity b/w u,courses
    sum1 = 0
    sum2 = 0
    sum3 = 0
    # avgV = getAvgGrades(sheet, v)
    # print(" avg ",avgU,avgV,v)
    for i in range(1,sheet.max_row):
        if IsScore(sheet,i,c1) and IsScore(sheet,i,c2):
            avg = getAvgGrades(sheet,i)
            # print(i,avg)
            sum1 = sum1 + (scr_sub(sheet,i,c1) - avg) * (scr_sub(sheet,i,c2) - avg)
            sum2 = sum2 + (scr_sub(sheet,i,c1) - avg) * (scr_sub(sheet,i,c1) - avg)
            sum3 = sum3 + (scr_sub(sheet,i,c2) - avg) * (scr_sub(sheet,i,c2) - avg)
    if sum2!=0 and sum3!=0:
        return round(float(sum1/((sum2*sum3)**(1/2))),3)
    else:
        return 0

def Reco(course,studid):
    similar_item_val = []
    simi_item = []
    avgU = getAvgGrades(sheet, studid)
    for i in range(2,sheet.max_column-1):
        if i!=course:
            sim = similarity(i,course,sheet)
            similar_item_val.append([course,i,sim])
        if sim > 0:
            simi_item.append([course,i,sim])

                           #predicting grade
    sum1 = 0
    sum2 = 0
    for i in range(0,len(simi_item)):
        if IsScore(sheet,studid,int(simi_item[i][1])):
            sum1 = sum1 + (simi_item[i][2] * scr_sub(sheet,studid,simi_item[i][1]))
            sum2 = sum2 + simi_item[i][2]
    if sum2 != 0:
        ress=str(round(float((sum1/sum2)/10),3))
        lb4.config(text=ress)
        return 0
def finalll():
    c = data.get()
    course = opt.index(c) + 1
    studid = int(tx1.get())
    Reco(course,studid)
def clrinput():
    tx1.delete(0,END)
    data.set("  CHOOSE  ")


lb1=Label(window,text="STUDENT ID : ",fg='blue',font=('Arial',15)).place(x=170,y=140,width=230)

data1=StringVar()
tx1=Entry(window,textvariable=data1,fg='blue',font=('Arial',15))
tx1.place(x=470,y=140,width=230)

lb2 = Label(window, text="CHOOSE A COURSE : ", fg='blue', font=('Arial', 15)).place(x=170,y=240,width=230)

lb3 = Label(window, text="RESULT : ", fg='blue', font=('Arial', 15)).place(x=170,y=340,width=230)

lb4 = Label(window, text="", fg='blue', font=('Arial', 15))
lb4.place(x=470,y=340,)

btn2 = Button(window, text="PREDICT GRADE :", fg='blue', font=('Arial', 14),command=finalll).place(x=155,y=440,width=260)  # ,command=rec_course


btn3 = Button(window, text="CLEAR INPUTS", fg='blue', font=('Arial', 14),command=clrinput).place(x=455,y=440,width=260)  # ,command=rec_course




window.mainloop()

