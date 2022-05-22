from tkinter import *
import os
import openpyxl

wb= openpyxl.load_workbook('data1/student_marks.xlsx')
wb1= openpyxl.load_workbook('data1/grade_result.xlsx')

sheetnamemarks = wb.get_sheet_names()
sheetnamecourse=wb1.get_sheet_names()

wsheet=wb.get_sheet_by_name(sheetnamemarks[0])
rsheet=wb1.get_sheet_by_name(sheetnamecourse[3])


opt = [
    "java", "python", "c#", "c", "c++", "networking", "operating system", "web technology", "software engineering",
    "internet of things", "data mining", "software project management", "data science", "english", "data structure",
    "artificial intelligence", "computer grapfics", "mathematics",
    "chemistry", "mobile computing", "natural language processing", "android development", "aws", "cyber security",
    "machine learning", "embeded system", "physics",
    "digital electronics", "image processing", "cloud computing", "linux", "oracle", "rdbms", "dbms", "statistics"
]

window = Tk()
window.title("COURSE RECOMMENDATION SYSTEM")
window.geometry('870x600')
lbA = Label(window, text="WELCOME  TO  COURSE  RECOMMENDATION  SYSTEM", fg='blue', font=('Arial', 20)).place(x=60,y=40)
window['bg'] = '#5d8a82'

data = StringVar()
data.set("  CHOOSE  ")
drop = OptionMenu(window, data, *opt).place(x=470,y=240,width=230)


#CODE

#print("Ranked List of Course based on evaluation criteria preferences of student")

#studid = int(input("Enter Student ID : "))

#course=input("enter the course : ")
def recc(studid,course):
    t = 0
    for i in range(2, 37):
        if (wsheet.cell(row=1, column=i).value == course):
            t = i
            #print(wsheet.cell(row=1, column=t).value)
    L1=[]
    min=t*3-4
    max=t*3-1
    for i in range (min,max):
        L1.append(round(rsheet.cell(row=studid+1,column=i).value,2))
    L1.sort()
    #print(L1)

    LZ=[]
    if(wsheet.cell(row=studid+1,column=t).value>90):
        #print("YOU ARE EXCELLENT YOU HAVE SCORED MORE THAN 90")
        LZ = []
        LZ.append(rsheet.cell(row=1,column=min).value)
        LZ.append(rsheet.cell(row=1,column=min+1).value)
        LZ.append(rsheet.cell(row=1,column=min+2).value)
    elif(wsheet.cell(row=studid+1,column=t).value - int(wsheet.cell(row=studid+1,column=38).value) < 0):
        for i in range (min,max):
            if(round(rsheet.cell(row=studid+1,column=i).value,2)==L1[len(L1)-1]):
                LZ = []
                LZ.append(rsheet.cell(row=1,column=i).value)

    elif (wsheet.cell(row=studid+1, column=t).value - int(wsheet.cell(row=studid+1, column=38).value) < 15):
        for i in range(min, max):
            if (round(rsheet.cell(row=studid + 1, column=i).value, 2) == L1[len(L1)-2]):
                LZ = []
                LZ.append(rsheet.cell(row=1, column=i).value)
    else:
        for i in range(min, max):
            if (round(rsheet.cell(row=studid + 1, column=i).value, 2) == L1[len(L1) - 3]):
                LZ = []
                LZ.append(rsheet.cell(row=1, column=i).value)
    lb4.config(text=str(LZ))
    return 0





def finalll():
    c = data.get()
    course = c     #opt.index(c) + 1
    studid = int(tx1.get())
    recc(studid,course)
def clrinput():
    data.set("CHOOSE")
    tx1.delete(0,END)


lb1=Label(window,text="STUDENT ID : ",fg='blue',font=('Arial',15)).place(x=170,y=140,width=230)

data1=StringVar()
tx1=Entry(window,textvariable=data1,fg='blue',font=('Arial',15))
tx1.place(x=470,y=140,width=230)

lb2 = Label(window, text="CHOOSE A COURSE : ", fg='blue', font=('Arial', 15)).place(x=170,y=240,width=230)

lb3 = Label(window, text="RESULT : ", fg='blue', font=('Arial', 15)).place(x=170,y=340,width=230)

lb4 = Label(window, text="", fg='blue', font=('Arial', 15))
lb4.place(x=470,y=340,)

btn2 = Button(window, text="RECOMMENDE COURSE :", fg='blue', font=('Arial', 14),command=finalll).place(x=155,y=440,width=260)  # ,command=rec_course


btn3 = Button(window, text="CLEAR INPUTS", fg='blue', font=('Arial', 14),command=clrinput).place(x=455,y=440,width=260)  # ,command=rec_course




window.mainloop()