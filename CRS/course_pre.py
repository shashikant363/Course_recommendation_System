import os
import openpyxl

wb= openpyxl.load_workbook('data1/student_marks.xlsx')
wb1= openpyxl.load_workbook('data1/grade_result.xlsx')


sheetnamemarks = wb.get_sheet_names() 
sheetnamecourse=wb1.get_sheet_names()


wsheet=wb.get_sheet_by_name(sheetnamemarks[0])
rsheet=wb1.get_sheet_by_name(sheetnamecourse[3])

print("Ranked List of Course based on evaluation criteria preferences of student")

studid = int(input("Enter Student ID : "))
for i in range (2,37):
    if (i-2)%9 == 0:
        print()
    print(wsheet.cell(row=1,column=i).value,end=" , ")
print()
course=input("enter the course : ")
t=0
for i in range (2,37):
    if(wsheet.cell(row=1,column=i).value==course):
        t=i
        print(wsheet.cell(row=1,column=t).value)
L1=[]
min=t*3-4
max=t*3-1
for i in range (min,max):
    L1.append(round(rsheet.cell(row=studid+1,column=i).value,2))
L1.sort()
print(L1)


if(wsheet.cell(row=studid+1,column=t).value>90):
    print("YOU ARE EXCELLENT YOU HAVE SCORED MORE THAN 90")
    print("YOU CAN CHOOSE ANY COURSE FROM : ",rsheet.cell(row=1,column=min).value,rsheet.cell(row=1,column=min+1).value,rsheet.cell(row=1,column=min+2).value)
elif(wsheet.cell(row=studid+1,column=t).value - int(wsheet.cell(row=studid+1,column=38).value) < 0):
    for i in range (min,max):
        if(round(rsheet.cell(row=studid+1,column=i).value,2)==L1[len(L1)-1]):
            print("RECOMENDED COURSE IS : ",rsheet.cell(row=1,column=i).value)

elif (wsheet.cell(row=studid+1, column=t).value - int(wsheet.cell(row=studid+1, column=38).value) < 15):
    for i in range(min, max):
        if (round(rsheet.cell(row=studid + 1, column=i).value, 2) == L1[len(L1)-2]):
            print("RECOMENDED COURSE IS : ",rsheet.cell(row=1, column=i).value)
else:
    for i in range(min, max):
        if (round(rsheet.cell(row=studid + 1, column=i).value, 2) == L1[len(L1) - 3]):
            print("RECOMENDED COURSE IS : ",rsheet.cell(row=1, column=i).value)


print("DONE")
